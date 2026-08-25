from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

FONT="Arial"
TITLE=Font(name=FONT,bold=True,size=14)
SUB=Font(name=FONT,italic=True,size=9,color="555555")
HDR=Font(name=FONT,bold=True,color="FFFFFF",size=10)
BOLD=Font(name=FONT,bold=True,size=10)
REG=Font(name=FONT,size=10)
INPUT=Font(name=FONT,size=10,color="0000FF")
NAVY=PatternFill("solid",fgColor="1F3864"); BLUE=PatternFill("solid",fgColor="2E5496")
GREEN=PatternFill("solid",fgColor="375623"); GREY=PatternFill("solid",fgColor="F2F2F2")
YEL=PatternFill("solid",fgColor="FFF2CC")
thin=Side(style="thin",color="BFBFBF"); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
CTR=Alignment(horizontal="center",vertical="center",wrap_text=True)
LFT=Alignment(horizontal="left",vertical="center",wrap_text=True)
USD='"$"#,##0.00'; PCT='0%'; PCT1='0.0%'
REDR=lambda rng,ws: ws.conditional_formatting.add(rng,CellIsRule(operator="lessThan",formula=["0"],fill=PatternFill("solid",fgColor="FFC7CE"),font=Font(name=FONT,color="9C0006")))
GRNR=lambda rng,ws: ws.conditional_formatting.add(rng,CellIsRule(operator="greaterThanOrEqual",formula=["0"],fill=PatternFill("solid",fgColor="C6EFCE"),font=Font(name=FONT,color="006100")))

wb=Workbook()

# ---------------- Inputs ----------------
inp=wb.active; inp.title="Inputs"
inp.column_dimensions["A"].width=36; inp.column_dimensions["B"].width=14; inp.column_dimensions["C"].width=56
inp["A1"]="Pricing Model — Inputs"; inp["A1"].font=TITLE
inp["A2"]="Edit the highlighted (blue) cells. Everything else recalculates."; inp["A2"].font=SUB
rows=[
("Your hourly value (human time)",150,"$/hr — value of an hour of your time",USD),
("  → Value per minute","=B4/60","auto = hourly / 60",USD),
("Agent cost per case",1.00,"AI/API/compute to handle one case",USD),
("Payment processing fee",0.03,"card/Stripe ~3%",PCT1),
("Tooling — fixed monthly",200,"Atera + Loki/Hetzner share + misc",USD),
("Est. monthly case volume",50,"spreads fixed cost across cases",'0'),
("  → Fixed cost per case","=B8/B9","auto = fixed monthly / volume",USD),
("Default session length (min)",60,"baseline session duration",'0'),
("Phase 1 automation % (you-heavy)",0.10,"agent share while you supervise everything",PCT),
("Phase 2 automation % (agent-heavy)",0.89,"agent share after graduation",PCT),
("Monthly-plan utilization %",0.60,"share of INCLUDED sessions subscribers actually use (breakage lives here)",PCT),
]
r=4
for label,val,note,fmt in rows:
    inp.cell(r,1,label).font=BOLD if not label.startswith("  ") else REG; inp.cell(r,1).alignment=LFT
    c=inp.cell(r,2)
    if isinstance(val,str) and val.startswith("="): c.value=val; c.font=REG
    else: c.value=val; c.font=INPUT; c.fill=YEL
    c.number_format=fmt; c.alignment=CTR; c.border=BORD
    inp.cell(r,3,note).font=SUB; inp.cell(r,3).alignment=LFT
    r+=1
# map: B4 rate,B5 valmin,B6 agent,B7 fee,B8 fixedmo,B9 vol,B10 fixedcase,B11 slen,B12 p1,B13 p2,B14 util
VALMIN="Inputs!$B$5"; AGENT="Inputs!$B$6"; FEE="Inputs!$B$7"; FIXEDCASE="Inputs!$B$10"
P1="Inputs!$B$12"; P2="Inputs!$B$13"; UTIL="Inputs!$B$14"

# ---------------- Cost to Serve ----------------
cs=wb.create_sheet("Cost to Serve")
cs["A1"]="Cost to Serve a 60-min Session vs. Automation"; cs["A1"].font=TITLE
cs["A2"]="As the agent handles more, your time drops out and cost falls. This gap IS the overhead saving — real only once agents deliver."; cs["A2"].font=SUB
h=["Automation %","Human minutes","Human cost","Agent cost","Fixed/case","Cost to Serve"]; hr=4
for i,t in enumerate(h,1):
    x=cs.cell(hr,i,t); x.font=HDR; x.fill=BLUE; x.alignment=CTR; x.border=BORD
for i,w in enumerate([13,14,13,12,12,14],1): cs.column_dimensions[chr(64+i)].width=w
for k,lv in enumerate([0,.10,.25,.50,.75,.89,.95,1.0]):
    rr=hr+1+k
    cs.cell(rr,1,lv).number_format=PCT
    cs.cell(rr,2,f"=Inputs!$B$11*(1-A{rr})").number_format='0'
    cs.cell(rr,3,f"=B{rr}*{VALMIN}").number_format=USD
    cs.cell(rr,4,f"={AGENT}").number_format=USD
    cs.cell(rr,5,f"={FIXEDCASE}").number_format=USD
    cs.cell(rr,6,f"=C{rr}+D{rr}+E{rr}").number_format=USD
    for c in range(1,7): cs.cell(rr,c).border=BORD; cs.cell(rr,c).alignment=CTR
    if k%2:
        for c in range(1,7): cs.cell(rr,c).fill=GREY
cs.freeze_panes="A5"

# ---------------- Tier Margins (current, now with utilization) ----------------
tm=wb.create_sheet("Tier Margins (Current)")
tm["A1"]="Current Prices — Phase 1 vs Phase 2, adjusted for utilization"; tm["A1"].font=TITLE
tm["A2"]="Total cost now reflects sessions ACTUALLY used (sessions x utilization). Red margin = losing money."; tm["A2"].font=SUB
heads=["Tier","Price","Billing","Sessions /cycle","Util %","Session min",
"Cost/sess P1","Total cost P1","Fees","Margin $ P1","Margin % P1",
"Cost/sess P2","Total cost P2","Margin $ P2","Margin % P2","Break-even auto %"]; hr=4
for i,t in enumerate(heads,1):
    x=tm.cell(hr,i,t); x.font=HDR; x.fill=NAVY; x.alignment=CTR; x.border=BORD
for i,w in enumerate([22,9,10,9,8,8, 10,10,8,10,10, 10,10,10,10, 14],1): tm.column_dimensions[chr(64+i)].width=w
# name,price,billing,sessions,util_formula(use UTIL for monthly,1 for one-time),smin
tiers=[("One-time session",75,"one-time",1,"1",60),
("Monthly support",49,"monthly",2,UTIL,60),
("Family plan",79,"monthly",4,UTIL,60),
("AI-first self-serve (old idea)",25,"monthly",4,UTIL,30)]
for k,(name,price,bill,sess,util,smin) in enumerate(tiers):
    rr=hr+1+k
    tm.cell(rr,1,name).font=REG; tm.cell(rr,1).alignment=LFT
    tm.cell(rr,2,price).font=INPUT; tm.cell(rr,2).fill=YEL; tm.cell(rr,2).number_format=USD
    tm.cell(rr,3,bill).font=REG
    tm.cell(rr,4,sess).font=INPUT; tm.cell(rr,4).fill=YEL
    tm.cell(rr,5,f"={util}").number_format=PCT
    tm.cell(rr,6,smin).font=INPUT; tm.cell(rr,6).fill=YEL
    tm.cell(rr,7,f"=F{rr}*(1-{P1})*{VALMIN}+{AGENT}+{FIXEDCASE}").number_format=USD
    tm.cell(rr,8,f"=G{rr}*D{rr}*E{rr}").number_format=USD
    tm.cell(rr,9,f"=B{rr}*{FEE}").number_format=USD
    tm.cell(rr,10,f"=B{rr}-H{rr}-I{rr}").number_format=USD
    tm.cell(rr,11,f"=J{rr}/B{rr}").number_format=PCT
    tm.cell(rr,12,f"=F{rr}*(1-{P2})*{VALMIN}+{AGENT}+{FIXEDCASE}").number_format=USD
    tm.cell(rr,13,f"=L{rr}*D{rr}*E{rr}").number_format=USD
    tm.cell(rr,14,f"=B{rr}-M{rr}-I{rr}").number_format=USD
    tm.cell(rr,15,f"=N{rr}/B{rr}").number_format=PCT
    tm.cell(rr,16,f"=1-((B{rr}*(1-{FEE})-D{rr}*E{rr}*({AGENT}+{FIXEDCASE}))/(D{rr}*E{rr}*F{rr}*{VALMIN}))").number_format=PCT
    for c in range(1,17): tm.cell(rr,c).border=BORD
    for c in range(2,17): tm.cell(rr,c).alignment=CTR
last=hr+len(tiers)
for col in ["K","O"]: REDR(f"{col}{hr+1}:{col}{last}",tm); GRNR(f"{col}{hr+1}:{col}{last}",tm)
tm.conditional_formatting.add(f"P{hr+1}:P{last}",CellIsRule(operator="greaterThan",formula=["1"],fill=PatternFill("solid",fgColor="FFC7CE"),font=Font(name=FONT,color="9C0006")))
tm.freeze_panes="B5"

# ---------------- Proposed Tiers ----------------
pt=wb.create_sheet("Proposed Tiers")
pt["A1"]="Proposed Tier Structure — built for agent delivery"; pt["A1"].font=TITLE
pt["A2"]="Each tier priced for its real automation level. Agent-only entry tier is cheap because it IS low-overhead. Founding rate is explicitly temporary."; pt["A2"].font=SUB
heads=["Tier","Price","Billing","Sessions /cycle","Util %","Session min","Automation %",
"Cost/sess","Total cost","Fees","Margin $","Margin %","Notes"]; hr=4
for i,t in enumerate(heads,1):
    x=pt.cell(hr,i,t); x.font=HDR; x.fill=GREEN; x.alignment=CTR; x.border=BORD
for i,w in enumerate([26,9,10,9,8,8,11, 10,10,8,10,9, 40],1): pt.column_dimensions[chr(64+i)].width=w
# name,price,billing,sessions,util,smin,automation,note
P=[
("AI Chat (agent-only)",19,"monthly",8,"0.5",10,"1.0","Unlimited chat modeled as 8 agent touches/mo. Cheap because it's truly low-overhead. The on-ramp for people who can't pay more."),
("Essential (human-backed)",59,"monthly",2,UTIL,60,P2,"AI-first, certified human on escalation. Hold price, add 24/7 + unlimited quick questions."),
("Family",99,"monthly",4,UTIL,60,P2,"Repriced from $79 — $79 was underwater even fully automated. $99 covers worst-case, profits on breakage."),
("Per-session",99,"one-time",1,"1",60,P2,"Raised from $75. No-commitment rate should sit ABOVE the monthly per-session equivalent to push subscriptions."),
("Founding rate — Essential (6 mo)",39,"monthly",2,UTIL,60,P2,"TEMPORARY. First 50 clients, 6 months, then reverts to $59. Lets you 'pass savings' now without anchoring low."),
]
for k,(name,price,bill,sess,util,smin,auto,note) in enumerate(P):
    rr=hr+1+k
    pt.cell(rr,1,name).font=REG; pt.cell(rr,1).alignment=LFT
    pt.cell(rr,2,price).font=INPUT; pt.cell(rr,2).fill=YEL; pt.cell(rr,2).number_format=USD
    pt.cell(rr,3,bill).font=REG
    pt.cell(rr,4,sess).font=INPUT; pt.cell(rr,4).fill=YEL
    pt.cell(rr,5,f"={util}").number_format=PCT
    pt.cell(rr,6,smin).font=INPUT; pt.cell(rr,6).fill=YEL
    pt.cell(rr,7,f"={auto}").number_format=PCT
    pt.cell(rr,8,f"=F{rr}*(1-G{rr})*{VALMIN}+{AGENT}+{FIXEDCASE}").number_format=USD
    pt.cell(rr,9,f"=H{rr}*D{rr}*E{rr}").number_format=USD
    pt.cell(rr,10,f"=B{rr}*{FEE}").number_format=USD
    pt.cell(rr,11,f"=B{rr}-I{rr}-J{rr}").number_format=USD
    pt.cell(rr,12,f"=K{rr}/B{rr}").number_format=PCT
    pt.cell(rr,13,note).font=SUB; pt.cell(rr,13).alignment=LFT
    for c in range(1,14): pt.cell(rr,c).border=BORD
    for c in range(2,13): pt.cell(rr,c).alignment=CTR
last=hr+len(P)
REDR(f"L{hr+1}:L{last}",pt); GRNR(f"L{hr+1}:L{last}",pt)
pt.freeze_panes="B5"

# ---------------- Read Me ----------------
rm=wb.create_sheet("Read Me")
rm.column_dimensions["A"].width=22; rm.column_dimensions["B"].width=92
notes=[
("Pricing Model — how to read it","T"),
("Inputs","Change blue cells. Hourly value, agent cost, fees, tooling, volume, two automation levels, and utilization drive everything."),
("Cost to Serve","How cost falls as the agent handles more. The high-vs-low-automation gap is your real overhead saving."),
("Tier Margins (Current)","Your live prices at Phase 1 (you deliver) vs Phase 2 (agent delivers), now adjusted for utilization. Red = loss."),
("Proposed Tiers","A structure that actually works: agent-only $19 on-ramp, human-backed Essential $59, repriced Family $99, $99 per-session, and a temporary $39 founding rate."),
("",""),
("KEY INSIGHT","K"),
("","At $150/hr a 60-min human session costs ~$150 to deliver. Consumer prices are PHASE-2 products — profitable only when the agent delivers. In Phase 1 they're a deliberate, time-boxed loss leader. Cap the volume."),
("UTILIZATION","K"),
("","Monthly margins depend heavily on how many included sessions go unused (breakage). Default 60% utilization. Price to survive worst-case (100% use), profit on breakage — never the reverse."),
("PASSING SAVINGS","K"),
("","Pass savings ONLY as categories graduate (use the tracker). Hold price + add value where AI is cheap; cut headline price only on the genuinely agent-only tier; use the founding rate to give an early discount without anchoring low."),
]
i=1
for a,b in notes:
    if b=="": i+=1; continue
    ca=rm.cell(i,1,a)
    if b=="T": ca.font=TITLE
    elif b=="K": ca.font=BOLD
    else: ca.font=BOLD; cb=rm.cell(i,2,b); cb.font=REG; cb.alignment=LFT
    ca.alignment=LFT; i+=1
wb.move_sheet("Read Me",-(wb.sheetnames.index("Read Me")))
wb.save("Klaravex-Personal-Pricing-Model.xlsx")
print("saved")
