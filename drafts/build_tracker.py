from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

FONT="Arial"
HDR=Font(name=FONT,bold=True,color="FFFFFF",size=11)
BOLD=Font(name=FONT,bold=True,size=11)
REG=Font(name=FONT,size=10)
TITLE=Font(name=FONT,bold=True,size=14)
SUB=Font(name=FONT,italic=True,size=9,color="555555")
NAVY=PatternFill("solid",fgColor="1F3864")
BLUE=PatternFill("solid",fgColor="2E5496")
GREENF=PatternFill("solid",fgColor="C6EFCE")
YELLOWF=PatternFill("solid",fgColor="FFEB9C")
REDF=PatternFill("solid",fgColor="FFC7CE")
GREY=PatternFill("solid",fgColor="F2F2F2")
thin=Side(style="thin",color="BFBFBF")
BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
CTR=Alignment(horizontal="center",vertical="center",wrap_text=True)
LFT=Alignment(horizontal="left",vertical="center",wrap_text=True)

cats=[("Password reset","Green"),("Wi-Fi / connectivity reconnect","Green"),
("App reinstall / update","Green"),("Printer / driver fix","Green"),
('"How do I" / usage question',"Green"),("Email / account setup","Yellow"),
("Software configuration","Yellow"),("Privacy / security settings","Yellow"),
("Smart-home device setup","Yellow"),("Data recovery","Red"),
("Security incident / malware","Red"),("Backup / restore","Red"),
("Payment / identity change","Red")]

wb=Workbook()

# ---------- Sheet 1: Case Log ----------
ws=wb.active; ws.title="Case Log"
ws["A1"]="Klaravex Personal — Case Log"; ws["A1"].font=TITLE
ws["A2"]="Log EVERY supervised case from day one. This is the data the Graduation Dashboard runs on — no log, no graduation."; ws["A2"].font=SUB
heads=["Case ID","Date","Customer","Category","Tier","Issue Summary","Handled By",
"Outcome","First-Attempt Correct? (Y/N)","Harm Event? (Y/N)","Escalated? (Y/N)","Notes"]
r=4
for c,h in enumerate(heads,1):
    cell=ws.cell(r,c,h); cell.font=HDR; cell.fill=NAVY; cell.alignment=CTR; cell.border=BORD
widths=[10,12,16,26,9,34,18,14,14,13,13,30]
for c,w in enumerate(widths,1):
    ws.column_dimensions[chr(64+c)].width=w
# blank rows
NROWS=60
for i in range(NROWS):
    rr=r+1+i
    for c in range(1,13):
        cell=ws.cell(rr,c); cell.border=BORD; cell.font=REG
        cell.alignment=LFT if c in (6,12) else CTR
    if i%2: 
        for c in range(1,13): ws.cell(rr,c).fill=GREY

# data validations
def dv(formula,col,allow_blank=True):
    d=DataValidation(type="list",formula1=formula,allow_blank=allow_blank)
    ws.add_data_validation(d)
    d.add(f"{col}{r+1}:{col}{r+NROWS}")
catlist='"'+",".join([c[0].replace('"',"") for c in cats])+'"'
# category list too long for inline (has commas/quotes) -> use reference range on Dashboard
dv('"Green,Yellow,Red"',"E")
dv('"You,Agent (supervised),Agent (autonomous)"',"G")
dv('"Resolved,Rework needed,Escalated"',"H")
dv('"Y,N"',"I"); dv('"Y,N"',"J"); dv('"Y,N"',"K")
ws.freeze_panes="A5"

# ---------- Sheet 2: Graduation Dashboard ----------
gd=wb.create_sheet("Graduation Dashboard")
gd["A1"]="Graduation Dashboard — Supervised → Autonomous"; gd["A1"].font=TITLE
gd["A2"]="Auto-calculated from Case Log. A category graduates only when Volume≥30, Accuracy≥95%, and Harm=0 — Red tier never graduates."; gd["A2"].font=SUB
gh=["Category","Tier","Cases Logged","First-Attempt OK","Accuracy %","Harm Events","Escalations","Graduation Status"]
hr=4
for c,h in enumerate(gh,1):
    cell=gd.cell(hr,c,h); cell.font=HDR; cell.fill=BLUE; cell.alignment=CTR; cell.border=BORD
gw=[28,9,13,14,12,12,12,30]
for c,w in enumerate(gw,1): gd.column_dimensions[chr(64+c)].width=w
CL="'Case Log'!"
for i,(name,tier) in enumerate(cats):
    rr=hr+1+i
    gd.cell(rr,1,name).font=REG; gd.cell(rr,1).alignment=LFT
    t=gd.cell(rr,2,tier); t.font=BOLD; t.alignment=CTR
    t.fill=GREENF if tier=="Green" else YELLOWF if tier=="Yellow" else REDF
    gd.cell(rr,3,f'=COUNTIF({CL}$D:$D,$A{rr})')
    gd.cell(rr,4,f'=COUNTIFS({CL}$D:$D,$A{rr},{CL}$I:$I,"Y")')
    gd.cell(rr,5,f'=IF(C{rr}=0,"",D{rr}/C{rr})').number_format="0.0%"
    gd.cell(rr,6,f'=COUNTIFS({CL}$D:$D,$A{rr},{CL}$J:$J,"Y")')
    gd.cell(rr,7,f'=COUNTIFS({CL}$D:$D,$A{rr},{CL}$K:$K,"Y")')
    status=(f'=IF($B{rr}="Red","HUMAN-ONLY (never graduates)",'
            f'IF(C{rr}<30,"BUILDING ("&C{rr}&"/30 cases)",'
            f'IF(AND(E{rr}>=0.95,F{rr}=0),'
            f'IF($B{rr}="Yellow","ELIGIBLE — approve-before-send","✓ ELIGIBLE — full autonomy"),'
            f'"HOLD — review (acc<95% or harm>0)")))')
    gd.cell(rr,8,status)
    for c in range(1,9):
        gd.cell(rr,c).border=BORD
        if c>=3: gd.cell(rr,c).alignment=CTR; gd.cell(rr,c).font=REG
gd.freeze_panes="A5"
# totals row
tr=hr+1+len(cats)
gd.cell(tr,1,"TOTAL").font=BOLD
gd.cell(tr,3,f'=SUM(C{hr+1}:C{tr-1})').font=BOLD
gd.cell(tr,6,f'=SUM(F{hr+1}:F{tr-1})').font=BOLD
gd.cell(tr,7,f'=SUM(G{hr+1}:G{tr-1})').font=BOLD
for c in range(1,9): gd.cell(tr,c).fill=GREY; gd.cell(tr,c).border=BORD; gd.cell(tr,c).alignment=CTR

# ---------- Sheet 3: How to use ----------
hw=wb.create_sheet("How to Use + Tiers")
hw.column_dimensions["A"].width=22; hw.column_dimensions["B"].width=85
rows=[
("Klaravex Personal — Trust & Graduation Framework","",TITLE),
("","",None),
("PURPOSE","Track every supervised case so AI task-categories can graduate to autonomy on evidence, not gut feel. Prevents Phase 1 (you on every ticket) from becoming permanent.",None),
("","",None),
("DAILY","Log each case in 'Case Log'. Fill Category, Tier, Handled By, Outcome, and the three Y/N flags. 30 seconds per case.",None),
("WEEKLY","Open 'Graduation Dashboard'. Look for categories nearing 30 cases and check accuracy + harm. Graduate eligible Green categories; review HOLDs.",None),
("","",None),
("GRADUATION GATE","All must be true to graduate a category:",BOLD),
("  Volume","≥ 30 supervised cases logged in that category",None),
("  Accuracy","≥ 95% first-attempt-correct (First-Attempt Correct = Y)",None),
("  Zero harm","0 harm events (no data loss, security exposure, or rework-causing damage)",None),
("  Escalation works","Agent reliably hands off its own edge cases (judge from the log)",None),
("","",None),
("TIERS","",BOLD),
("  Green","Graduate first. Reversible, low-stakes: password resets, Wi-Fi, app reinstall, printer, usage questions. → full autonomy once gate met.",None),
("  Yellow","Graduate late, with guardrails: email/account setup, software config, privacy settings, smart-home. → agent drafts, human approves before it reaches the customer.",None),
("  Red","NEVER fully autonomous: data recovery, security incidents, backups, payments/identity, vulnerable customers. Human-in-loop permanently, regardless of track record.",None),
("","",None),
("REVOCABLE","Any graduated category that produces a harm event drops back to supervised immediately. Trust is revocable.",None),
("DISCLOSURE LINK","The moment ANY category goes autonomous, the website must switch to Phase-2 AI-delivery disclosure — even if most work is still supervised. See personal-AI-disclosure-DRAFT.md.",None),
("","",None),
("THRESHOLDS","30 / 95% / 0-harm are starting points. Tighten for higher-risk categories; never loosen below these for anything customer-facing.",SUB),
]
for i,(a,b,f) in enumerate(rows,1):
    ca=hw.cell(i,1,a); cb=hw.cell(i,2,b)
    if f: ca.font=f
    else: ca.font=BOLD if a.strip() and not a.startswith("  ") else REG
    cb.font=REG; cb.alignment=LFT; ca.alignment=LFT
    if a.startswith("  "): ca.font=BOLD

# order sheets: How to Use first
wb.move_sheet("How to Use + Tiers",-(wb.sheetnames.index("How to Use + Tiers")))
wb.save("Klaravex-Personal-Case-Log-Graduation-Tracker.xlsx")
print("saved")
